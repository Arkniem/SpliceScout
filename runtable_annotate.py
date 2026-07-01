# -*- coding: utf-8 -*-
"""
Deep-dive stage — ANNOTATE: add drug / dose / is_control to the FILTERED run table, then build a
filterable Excel workbook.

Drug/control use the MAIN pipeline's general canonicalization (normalize_v2 + compound_map.json) so
this generalizes to any query — NOT the second pipeline's hardcoded VOCAB (specific to its 48-study
set). Dose uses the second pipeline's general DOSE regex (the one piece worth keeping). Writes a
drug_annotation_review.csv audit and a per-cell-line .xlsx (openpyxl; skipped gracefully if absent).
"""
import os
import re
import csv
import json

from normalize_v2 import clean_compound, is_control as nv_is_control
from build_final import _safe_open
from cellline_match import _slug
from progress import NULL

# treatment-ish columns (run-table attribute tags), in priority order
TREATMENT_COLS = ["treatment", "treatments", "agent", "agents", "compound", "compounds",
                  "drug", "drugs", "drug_treatment", "treated_with", "chemical",
                  "perturbation", "small_molecule", "inhibitor", "stimulus"]

DOSE = re.compile(r"(\d+(?:\.\d+)?)\s*(nM|µM|μM|uM|mM|ng/?mL|µg/?mL|μg/?mL|ug/?mL|mg/?mL|%|Gy)\b", re.I)


def parse_dose(text):
    out = "; ".join(dict.fromkeys("%s %s" % (m.group(1), m.group(2)) for m in DOSE.finditer(text or "")))
    m = re.search(r"(water|ethanol)\s+extract", text or "", re.I)
    if m:
        out = (out + " " if out else "") + m.group(1).lower() + " extract"
    return out.strip()


def _disp(name):
    return " + ".join((p.strip()[:1].upper() + p.strip()[1:])
                      for p in str(name).split("+") if p.strip())


_NONDRUG_INPUT = re.compile(r"(total|polya|polyadenylated|genomic|plasmid)?"
                            r"(sg|si|sh|g|m|t|r|nc|sn|sno|mi|lnc|ss|ds|c|pre|mt|cf)?(rna|dna)")


def _is_nondrug_input(token):
    """A nucleic-acid INPUT or sequencing-method token (RNA/DNA/STRT/ERCC/spike-in) — NEVER a drug, even with a
    dose attached. Stops an RNA/DNA input AMOUNT (e.g. 'STRT 10ug', '10ug total RNA', 'cDNA') from being read as
    a drug dose -> a false 'Drug Treated' (canon_drug otherwise treats ANY leftover clean token as a compound).
    Hit LIVE: K562 GSM3057932-39 'STRT 10ug … RNA-Seq' = an RNA-input titration, not a drug treatment."""
    t = re.sub(r"[^a-z0-9]", "", (token or "").lower())
    return t in {"strt", "ercc", "erccspikein", "spikein", "input"} or bool(_NONDRUG_INPUT.fullmatch(t))


def canon_drug(raw, compound_map):
    """Canonical generic drug name (AI map when present), or '' for controls / non-drugs."""
    c = clean_compound(raw)               # None for vehicle/control/empty
    if not c:
        return ""
    if _is_nondrug_input(c):              # RNA/DNA input or seq-method token -> not a drug (see _is_nondrug_input)
        return ""
    info = compound_map.get(c)
    if info is None:
        return _disp(c)
    if not info.get("is_drug", True):
        return ""
    return _disp(info.get("name") or c)


def drug_treated_label(raw, drug, compound_map):
    """3-way per-run classification: Drug Treated / Not Drug Treated / Undetermined.

    Real drug -> treated; explicit control or a recognized non-drug perturbation (is_drug=False,
    e.g. siRNA/KO) -> not treated; no treatment value or a present-but-unrecognized one -> undetermined.
    """
    if drug:
        return "Drug Treated"
    if not raw:
        return "Undetermined"
    if nv_is_control(raw):
        return "Not Drug Treated"
    c = clean_compound(raw)
    info = compound_map.get(c) if c else None
    if info is not None and not info.get("is_drug", True):
        return "Not Drug Treated"     # known non-drug perturbation
    return "Undetermined"             # present but can't be classified (e.g. skip-AI, novel term)


_TIME = re.compile(r"\b\d+(?:\.\d+)?\s*(?:h|hr|hrs|hour|hours|d|day|days|min|mins|minute|minutes|"
                   r"wk|wks|week|weeks)\b", re.I)
_FRAMING = re.compile(r"\b(?:treatment|treated|treat|treating|exposure|exposed|stimulation|stimulated|"
                      r"incubation|incubated|administration|administered|dosed?|application|applied|"
                      r"for|with|at)\b", re.I)


def _core_treatment(raw):
    """Strip dose/time/framing noise so a clear control or compound resurfaces, e.g.
    'PBS treatment' -> 'PBS', 'vehicle treated' -> 'vehicle', '12 h, 5uM BRM014' -> 'BRM014'.
    The control-detection and compound vocab match the bare token but MISS it under this framing, which is
    why ~a third of A549 runs came back 'Undetermined' despite a clear treatment string."""
    s = DOSE.sub(" ", raw or "")
    s = _TIME.sub(" ", s)
    s = _FRAMING.sub(" ", s)
    s = re.sub(r"[,;:]+", " ", s)
    return re.sub(r"\s+", " ", s).strip(" -")


def recover_label(raw, compound_map, drug_names):
    """Rescue a run the 3-way label left 'Undetermined' by re-testing its NOISE-STRIPPED core. Returns
    (label, drug) where label is 'Drug Treated'/'Not Drug Treated' (or None to stay Undetermined) and drug is
    the recovered canonical name (or ''). CONSERVATIVE on the drug side — a compound flips to Drug Treated
    ONLY if the AI compound map confirms it (exact raw key, or its canonical name appears as a confirmed
    is_drug=True elsewhere). Mislabeling contaminates the treated-vs-control PSI comparison, whereas a dropped
    Undetermined run is harmless, so a novel/ambiguous token (e.g. an unrecognized perturbation) stays
    Undetermined rather than being guessed as a drug."""
    core = _core_treatment(raw)
    if not core or core.lower() == (raw or "").strip().lower():
        return None, ""               # nothing was stripped -> the base logic already saw this exact string
    if nv_is_control(core):
        return "Not Drug Treated", ""
    c = clean_compound(core)
    if c:
        info = compound_map.get(c)
        if info is not None:
            return ("Drug Treated", _disp(info.get("name") or c)) if info.get("is_drug", True) \
                   else ("Not Drug Treated", "")
        if c.lower() in drug_names:    # canonical name confirmed as a drug elsewhere in the map
            return "Drug Treated", _disp(c)
    return None, ""


def _treatment_value(row, cols_present):
    for c in cols_present:
        v = (row.get(c) or "").strip()
        if v:
            return v
    return ""


def _find_filtered_csv(P, slug):
    for cand in (P.runtable_filtered_csv(slug), P.runtable_filtered_csv(slug).replace(".csv", "_v2.csv")):
        if os.path.exists(cand):
            return cand
    return None


# A workbook can't exceed Excel's 1,048,576-row sheet limit, and a near-limit .xlsx is multi-GB even streamed
# (this K562 set's 1.15M-row table overran a 13 GB disk as write-only XML -> ENOSPC). Above this cap we skip
# the workbook and leave the CSV as the full deliverable. Override with RUNTABLE_XLSX_MAX_ROWS.
XLSX_MAX_ROWS = max(1, int(os.environ.get("RUNTABLE_XLSX_MAX_ROWS", "100000") or "100000"))


def make_workbook(src_csv, out_xlsx, hdr=None, data=None):
    """Filterable workbook (All runs + Study summary). Ported from 04_make_workbook.py.
    Skips (returns None) above XLSX_MAX_ROWS -- Excel can't hold it and the streamed XML can blow the disk."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.cell import WriteOnlyCell
    except Exception as e:
        print(f"  WORKBOOK: openpyxl unavailable ({e}) -> skipping .xlsx")
        return None
    NUMERIC = {"AvgSpotLen", "Bases", "Bytes", "Spots", "TaxID", "version"}
    # PERF: run() passes the already-materialized table (hdr+data) so we skip a full re-read+parse of the CSV
    # we just wrote. Standalone callers (no hdr/data) fall back to reading src_csv. Values are identical
    # strings either way (csv round-trips strings losslessly), so the workbook is byte-for-byte the same.
    if hdr is None or data is None:
        with open(src_csv, encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if not rows:
            return None
        hdr, data = rows[0], rows[1:]
    if not hdr:
        return None
    if len(data) > XLSX_MAX_ROWS:                      # too big for Excel / the disk -> the CSV is the deliverable
        print(f"  WORKBOOK: {len(data):,} rows exceeds the {XLSX_MAX_ROWS:,}-row cap (Excel max 1,048,576) "
              f"-> skipping .xlsx; use {os.path.basename(src_csv)}.")
        return None
    numidx = {i for i, h in enumerate(hdr) if h in NUMERIC}
    fill = PatternFill("solid", fgColor="1F4E78"); font = Font(bold=True, color="FFFFFF")

    def _hdr_cells(ws, names):                     # styled header cells (write-only has no post-hoc cell access)
        cells = []
        for v in names:
            c = WriteOnlyCell(ws, value=v); c.fill = fill; c.font = font
            cells.append(c)
        return cells

    # WRITE-ONLY workbook: rows stream straight to the .xlsx at ~constant memory instead of materializing a
    # Cell object per cell. A filtered table of 8k rows x ~350 cols is millions of cells -> openpyxl's normal
    # mode OOM'd on a 16 GB box. Same two sheets / styling / freeze / filter, just streamed.
    wb = openpyxl.Workbook(write_only=True)

    ws = wb.create_sheet("All runs")
    ws.freeze_panes = "B2"
    for i, h in enumerate(hdr, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(max(len(h) + 2, 11), 32)
    ws.append(_hdr_cells(ws, hdr))
    nrows = 0
    for r in data:
        ws.append([int(v) if i in numidx and v.strip().lstrip("-").isdigit() else v
                   for i, v in enumerate(r)])
        nrows += 1
    if nrows:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(hdr)), nrows + 1)

    ws2 = wb.create_sheet("Study summary")
    ws2.freeze_panes = "B2"
    gi = hdr.index("GSE_Series") if "GSE_Series" in hdr else None
    if gi is not None:
        cols = {c: hdr.index(c) for c in ("SRA Study", "BioProject", "Organism", "Assay Type",
                                          "Platform", "Instrument") if c in hdr}
        head2 = ["GSE"] + list(cols) + ["# Runs"]
        ws2.append(_hdr_cells(ws2, head2))
        from collections import OrderedDict, Counter
        seen = OrderedDict(); cnt = Counter()
        for r in data:
            g = r[gi]; cnt[g] += 1
            if g not in seen:
                seen[g] = [r[cols[c]] for c in cols]
        for g, vals in seen.items():
            ws2.append([g] + vals + [cnt[g]])
        if seen:
            ws2.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(head2)), len(seen) + 1)

    try:
        wb.save(out_xlsx)
    except PermissionError:
        out_xlsx = out_xlsx.replace(".xlsx", "_v2.xlsx")
        wb.save(out_xlsx)
        print(f"  ** workbook locked -> wrote {os.path.basename(out_xlsx)} **")
    print(f"  WORKBOOK -> {out_xlsx} ({len(data)} runs)")
    return out_xlsx


def run(P, sel, reporter=NULL):
    """Annotate the filtered run table (drug/dose/is_control), write review + workbook."""
    slug = _slug(sel.get("canonical", "cellline"))
    src = _find_filtered_csv(P, slug)
    if not src:
        # consistent return contract: always a dict; signal the skip via n_runs=0 + skipped flag
        print(f"  ANNOTATE: filtered run table not found for {slug!r} "
              f"(looked for {P.runtable_filtered_csv(slug)} and its _v2 variant) -> skipping")
        return {"n_runs": 0, "is_control": {"yes": 0, "no": 0},
                "drug_treated": {"Drug Treated": 0, "Not Drug Treated": 0, "Undetermined": 0},
                "workbook": False, "annotated_csv": None, "skipped": True}
    compound_map = {}
    if os.path.exists(P.compound_map):
        try:
            with open(P.compound_map, encoding="utf-8") as f:
                compound_map = json.load(f)
        except Exception:
            compound_map = {}

    with open(src, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    base_cols = list(rows[0].keys()) if rows else []
    cols_present = [c for c in TREATMENT_COLS if c in base_cols]
    extra = ("drug", "dose", "is_control", "drug_treated")
    fields = [c for c in base_cols if c not in extra] + list(extra)

    # AI TITLE-based drug_treated as a FALLBACK for runs the treatment COLUMNS leave Undetermined — the exact
    # hybrid build_final uses for cellline_index (column treatment first, else sample_map.get(title)), so the
    # PSI grouping agrees with the headline counts instead of dropping samples the AI already classified. Keyed
    # by the GEO sample title from raw_json (build_final's key). Absent maps => fallback simply never fires.
    sample_map = {}
    if os.path.exists(P.sample_map):
        try:
            with open(P.sample_map, encoding="utf-8") as f:
                sample_map = json.load(f)
        except Exception:
            sample_map = {}
    gsm2title = {}
    if sample_map and os.path.exists(P.raw_json):
        try:
            res = json.load(open(P.raw_json, encoding="utf-8")).get("result", {})
            for u in res.get("uids", []):
                for s in (res.get(u, {}).get("samples") or []):
                    if s.get("accession"):
                        gsm2title[s["accession"]] = s.get("title", "")
        except Exception:
            gsm2title = {}
    gsm_col = next((c for c in base_cols if c.replace(" ", "").lower().startswith("geo_accession")),
                   "GEO_Accession (exp)")

    # canonical names the AI compound pass confirmed are real drugs -> lets recover_label flip a noise-framed
    # compound (e.g. '12 h, 5uM BRM014') to Drug Treated without ever guessing on an unrecognized token.
    drug_names = {(v.get("name") or "").strip().lower() for v in compound_map.values() if v.get("is_drug")}
    drug_names.discard("")

    review = {}
    counter = {"yes": 0, "no": 0}
    dt_counter = {"Drug Treated": 0, "Not Drug Treated": 0, "Undetermined": 0}
    recovered_noise = recovered_ai = 0
    # PERF: drug/dose/is_control/drug_treated are PURE functions of `raw` (+ the read-only compound_map), and
    # the same raw treatment string repeats across many runs (DMSO, untreated, a dose token, …). Memoize by
    # raw so the regex-heavy clean_compound/is_control/drug_treated_label run ONCE per distinct value, not per
    # run. Identical per-row assignment, counters, and review tallies — output unchanged.
    _ann, _rec = {}, set()
    for r in rows:
        raw = _treatment_value(r, cols_present)
        hit = _ann.get(raw)
        if hit is None:
            drug = canon_drug(raw, compound_map)
            dose = parse_dose(raw)
            # a real drug => not a control; else 'yes' only for explicit vehicle/negative controls
            # (a non-drug perturbation such as siRNA leaves drug='' but is_control='no').
            ctrl = "yes" if (not drug and raw and nv_is_control(raw)) else "no"
            dt = drug_treated_label(raw, drug, compound_map)   # 3-way Drug Treated/Not/Undetermined
            if dt == "Undetermined":                           # RECOVERY: re-test the noise-stripped core
                rlabel, rdrug = recover_label(raw, compound_map, drug_names)
                if rlabel:
                    dt = rlabel
                    _rec.add(raw)
                    if rdrug and not drug:
                        drug = rdrug
                    if rlabel == "Not Drug Treated" and ctrl == "no" and nv_is_control(_core_treatment(raw)):
                        ctrl = "yes"                           # a control the framing had hidden
            _ann[raw] = (drug, dose, ctrl, dt)
        else:
            drug, dose, ctrl, dt = hit
        if raw in _rec:
            recovered_noise += 1
        # review = the condition-based (column+noise) classification, per distinct raw treatment string
        _, _, _, _, n = review.get(raw, (None, None, None, None, 0))
        review[raw] = (drug, dose, ctrl, dt, n + 1)
        # FINAL label: for a run STILL Undetermined from its columns, defer to the AI TITLE classification
        # (per-GSM) — only a confident Drug/Not call is adopted (N/A & Undetermined stay Undetermined).
        if dt == "Undetermined" and gsm2title:
            adt = (sample_map.get(gsm2title.get((r.get(gsm_col) or "").strip(), "")) or {}).get("drug_treated", "")
            if adt in ("Drug Treated", "Not Drug Treated"):
                dt = adt
                recovered_ai += 1
        r["drug"], r["dose"], r["is_control"], r["drug_treated"] = drug, dose, ctrl, dt
        counter[ctrl] = counter.get(ctrl, 0) + 1
        dt_counter[dt] = dt_counter.get(dt, 0) + 1
    _rec_total = recovered_noise + recovered_ai
    reporter.set_detail(f"{len(rows)} runs annotated"
                        + (f" ({_rec_total} recovered from Undetermined)" if _rec_total else ""))

    out = _safe_open(src)  # rewrite in place (same file), _v2 fallback if locked
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore"); w.writeheader()
        for r in rows:
            w.writerow(r)
    with open(_safe_open(P.runtable_drug_review), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["original_condition", "drug", "dose", "is_control", "drug_treated", "n_runs"])
        for cond, (drug, dose, ctrl, dt, n) in sorted(review.items()):
            w.writerow([cond, drug, dose, ctrl, dt, n])

    # Skip the workbook (and its multi-GB in-memory copy) for oversized tables -- a 1.15M-row K562 set is past
    # Excel's ceiling and its write-only XML overran the disk. The CSV above is the full deliverable.
    if len(rows) > XLSX_MAX_ROWS:
        print(f"  WORKBOOK: {len(rows):,} rows exceeds the {XLSX_MAX_ROWS:,}-row cap (Excel max 1,048,576) "
              f"-> skipping .xlsx; the full annotated table is {os.path.basename(out)}.")
        xlsx = None
    else:
        # pass the in-memory table so make_workbook need not re-read+parse the CSV we just wrote
        _hdr = fields
        _data = [[str(r.get(h, "")) for h in fields] for r in rows]
        xlsx = make_workbook(out, P.runtable_workbook(slug), hdr=_hdr, data=_data)
    print(f"  ANNOTATE: {len(rows)} runs | drug_treated={dt_counter} | is_control={counter}"
          + (f" | recovered {_rec_total} from Undetermined (noise={recovered_noise}, AI-title={recovered_ai})"
             if _rec_total else "")
          + f" | review -> {os.path.basename(P.runtable_drug_review)}")
    return {"n_runs": len(rows), "is_control": counter, "drug_treated": dt_counter,
            "recovered": _rec_total, "recovered_noise": recovered_noise, "recovered_ai": recovered_ai,
            "workbook": bool(xlsx), "annotated_csv": out, "skipped": False}


def main():
    import argparse
    from pipeline_paths import Paths
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-dir", required=True)
    a = ap.parse_args()
    P = Paths(a.run_dir).ensure_dirs()
    with open(P.cellline_selection, encoding="utf-8") as f:
        sel = json.load(f)
    run(P, sel)


if __name__ == "__main__":
    main()
